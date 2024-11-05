from flask import Flask, render_template, request, flash, redirect, url_for
import openpyxl
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # necessary for flash messages

# Path to the Excel file
EXCEL_FILE_PATH = 'static/files/Credits and Debits Expen.xlsx'

def load_excel():
    """ Load the Excel file and return the active worksheet """
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    return workbook

def save_excel(workbook):
    """ Save changes to the Excel file """
    workbook.save(EXCEL_FILE_PATH)

@app.route('/')
def home():
    return render_template('main.html')

@app.route('/signup', methods=['POST'])
def signup():
    # Retrieve form data
    name = request.form['name']
    email = request.form['email']
    shift_date = request.form['shift-date']
    shift_time = request.form['shift-time']

    # Parse the shift time to calculate hours worked
    start_time, end_time = shift_time.split('-')
    start_hour = int(start_time.split(':')[0])
    end_hour = int(end_time.split(':')[0])
    hours_worked = end_hour - start_hour  # Calculate based on 17:00 as the closing time

    # Calculate points
    points_earned = hours_worked * 20

    # Load and update the Excel sheet
    workbook = load_excel()
    credits_sheet = workbook['Credits']
    # Parse the shift date
    shift_date_obj = datetime.strptime(shift_date, '%Y-%m-%d')

    # Check if the volunteer has worked recently
    for row in credits_sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is headers
        if row[0] == name and row[1] == email:
            last_shift_date = datetime.strptime(row[2], '%Y-%m-%d')
            if (shift_date_obj - last_shift_date).days < 7:
                flash("You need a 7-day gap between shifts. Please select a different date.")
                return redirect(url_for('home'))

    # Append new signup entry in Excel with hours worked and points earned
    credits_sheet.append([name, email, shift_date, shift_time, points_earned, 0, hours_worked])
    save_excel(workbook)
    flash("Sign-up successful! You will earn {} points for this shift.".format(points_earned))
    return redirect(url_for('home'))


@app.route('/confirm', methods=['POST'])
def confirm_attendance():
    name = request.form['name']
    shift_date = request.form['shift-date']

    # Verify that confirmation is allowed only during working hours
    current_hour = datetime.now().hour
    if not (9 <= current_hour <= 17):  # Modify hours if needed
        flash("Confirmations are only allowed between 9 AM and 5 PM.")
        return redirect(url_for('home'))

    # Load and update the Excel sheet for attendance
    workbook = load_excel()
    credits_sheet = workbook['Credits']
    confirmed = False

    # Search for the record in Excel to confirm attendance
    for row in credits_sheet.iter_rows(min_row=2):  # Assuming first row is headers
        if row[0].value == name and row[2].value == shift_date:
            row[4].value = 'Confirmed'  # Update attendance status
            confirmed = True
            break

    if confirmed:
        save_excel(workbook)
        flash("Attendance confirmed!")
    else:
        flash("No matching record found for confirmation.")
    
    return redirect(url_for('home'))


@app.route('/debit', methods=['POST'])
def debit():
    name = request.form['name']
    debit_amount = float(request.form['debit-amount'])  # Assuming it's a valid number
    debit_date = request.form['debit-date']

    # Load and update the Excel sheet for debits
    workbook = load_excel()
    credits_sheet = workbook['Credits']
    debits_sheet = workbook['Debits']
    
    # Find the volunteer's current balance
    current_balance = 0
    for row in credits_sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is headers
        if row[0] == name:
            current_balance = row[4]
            break

    # Check if the debit amount is valid (doesn't exceed the balance)
    if debit_amount <= current_balance:
        # Add the debit to the "Debits" sheet
        debits_sheet.append([name, debit_date, debit_amount])

        # Update balance in "Credits" sheet
        for row in credits_sheet.iter_rows(min_row=2):
            if row[0].value == name:
                row[5].value -= debit_amount
                break

        save_excel(workbook)
        flash("Debit of {} points successfully processed.".format(debit_amount))
    else:
        flash("Insufficient points to complete the debit.")
    
    return redirect(url_for('home'))


if __name__ == '__main__':
    app.run(debug=True)
